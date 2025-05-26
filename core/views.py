import decimal
import uuid
from collections import defaultdict
from datetime import timedelta

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Prefetch, Q, Sum
from django.http import Http404, HttpResponse, JsonResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl.utils import get_column_letter

from accounts.models import Account, Customer, Employee
from core.models import (
    Accessory,
    ItemSet,
    ItemSetComponent,
    ItemStatus,
    ItemType,
    Payment,
    PaymentStatus,
    PaymentType,
    RentalItem,
    RentalItemDetail,
    RentalItemType,
    RentalSetDetail,
    RentalTransaction,
)

from .forms import (
    AccessoryForm,
    BookingForm,
    CustomerForm,
    EmployeeForm,
    ItemSetComponentForm,
    ItemSetForm,
    RentalItemForm,
    RentalItemTypeForm,
    ReturnSettlementForm,
)


# --- Client Views ---
def custom_404_view(request, exception):
    return render(request, "404.html", status=404)


# NOTE: Home
def home(request):
    rental_items = RentalItem.objects.filter(status=ItemStatus.AVAILABLE).order_by("-updated_at")
    item_sets = ItemSet.objects.all().order_by("name")  # Consider adding availability check later

    # Calculate available quantity for sets (can be slow if many sets/items)
    # You might want to optimize this or calculate on demand
    available_sets = []
    for item_set in item_sets:
        if item_set.available_quantity > 0:
            available_sets.append(item_set)
            print("Available sets - inside: ", available_sets)
    print("Available sets - outside: ", available_sets)

    context = {
        "rental_items": rental_items,
        "item_sets": item_sets,
    }
    return render(request, "core/clients/pages/home.html", context)


def about(request):
    return render(request, "core/clients/pages/about.html")


def contact(request):
    return render(request, "core/clients/pages/contact.html")


def item_detail_view(request, pk):
    # Can show details for RentalItem or ItemSet based on a type parameter or separate URLs
    # For now, assuming it's for RentalItem
    item = get_object_or_404(RentalItem, pk=pk)
    context = {"item": item}
    return render(request, "core/clients/pages/item_detail.html", context)


@login_required  # Ensure only logged-in users can book
def create_booking_view(request, item_pk):
    item_to_book = get_object_or_404(RentalItem, pk=item_pk)

    # 1. Check if item is available
    if item_to_book.status != ItemStatus.AVAILABLE:
        messages.error(request, f"Sorry, '{item_to_book.item_type}' is currently unavailable.")
        # Redirect back to home or item detail page
        return redirect(reverse("home"))  # Or 'item_detail', item_to_book.pk

    # 2. Handle Form Submission (POST request)
    if request.method == "POST":
        form = BookingForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]
            payment_method = form.cleaned_data["payment_method"]

            # --- Data Validation & Calculation ---
            # Check for overlapping bookings for THIS specific item (simple check)
            overlapping_rentals = RentalItemDetail.objects.filter(
                item=item_to_book,
                rental__end_date__gte=start_date,  # Rental ends on or after new start
                rental__start_date__lte=end_date,  # Rental starts on or before new end
            ).exclude(
                rental__payment_status=PaymentStatus.CANCELLED
            )  # Optional: exclude cancelled

            if overlapping_rentals.exists():
                messages.error(
                    request,
                    f"'{item_to_book.item_type}' is already booked during the selected dates. Please choose different dates.",
                )
                # Re-render form with error implicitly handled by messages framework
                context = {"form": form, "item": item_to_book}
                return render(request, "core/clients/pages/booking_form.html", context)

            # Calculate rental duration (inclusive of start and end day)
            duration_days = (end_date - start_date).days + 1
            if duration_days <= 0:
                duration_days = 1  # Minimum 1 day rental

            # Calculate total rental cost
            price_per_day = item_to_book.item_type.rental_price_per_day
            total_rental_cost = price_per_day * decimal.Decimal(duration_days)

            # Calculate deposit (Example: 20% of total rental or minimum $20)
            deposit_percentage = decimal.Decimal("0.20")
            calculated_deposit = total_rental_cost * deposit_percentage
            minimum_deposit = decimal.Decimal("20.00")
            total_deposit = max(calculated_deposit, minimum_deposit)

            # --- Database Operations (Use a transaction) ---
            try:
                with transaction.atomic():
                    # Get the customer profile associated with the logged-in user
                    customer = get_object_or_404(Customer, user=request.user)

                    # a. Create RentalTransaction
                    rental_transaction = RentalTransaction.objects.create(
                        customer=customer,
                        start_date=start_date,
                        end_date=end_date,
                        total_deposit=total_deposit,
                        # total_fines starts at 0 (default)
                        payment_status=PaymentStatus.PENDING,  # Start as pending until payment confirmed
                    )

                    # b. Create RentalItemDetail
                    RentalItemDetail.objects.create(
                        rental=rental_transaction,
                        item=item_to_book,
                        quantity=1,  # Booking one specific item instance
                        rented_price_per_day=price_per_day,
                    )

                    # c. Create Initial Payment Record (for the deposit)
                    # In a real app, you'd integrate with a payment gateway here.
                    # For now, we'll create a PENDING payment record.
                    # You might change payment_status on RentalTransaction to PAID
                    # if payment is processed immediately and successfully.
                    payment = Payment.objects.create(
                        rental=rental_transaction,
                        amount=total_deposit,  # Paying the deposit amount initially
                        payment_method=payment_method,
                        payment_type=PaymentType.DEPOSIT,
                        # damage=None (no damage yet)
                        # transaction_date is auto_now_add
                        notes=f"Initial deposit for Rental #{rental_transaction.id}",
                    )

                    # d. Update Item Status (Optional: could be RENTED or PENDING_RENTAL)
                    # Let's set it to RENTED assuming the booking is confirmed
                    item_to_book.status = ItemStatus.RENTED
                    item_to_book.save()

                # --- Success ---
                messages.success(
                    request,
                    f"Booking for '{item_to_book.item_type}' created successfully! Please complete payment if required.",
                )
                # Redirect to a success page showing booking details
                return redirect(
                    reverse("booking_success", kwargs={"transaction_pk": rental_transaction.pk})
                )

            except Exception as e:
                # Handle potential errors during transaction
                messages.error(request, f"An error occurred while creating the booking: {e}")
                # Fall through to re-render the form

    # 3. Handle Initial Page Load (GET request)
    else:
        form = BookingForm()  # Create an empty form

    context = {"form": form, "item": item_to_book}
    return render(request, "core/clients/pages/booking_form.html", context)


# --- Helper Function for Session Management ---
def get_booking_selection(session):
    """Gets the booking selection from the session, initializing if needed."""
    selection = session.get("booking_selection", {"items": {}, "sets": {}})
    # Ensure keys exist
    if "items" not in selection:
        selection["items"] = {}
    if "sets" not in selection:
        selection["sets"] = {}
    return selection


def clear_booking_selection(session):
    """Clears the booking selection from the session."""
    if "booking_selection" in session:
        del session["booking_selection"]


# --- Booking Selection (Cart) Views ---


@require_POST  # Ensure this view only accepts POST requests
@login_required  # Require login to add items
def add_to_booking_selection(request):
    item_id = request.POST.get("item_id")
    set_id = request.POST.get("set_id")
    quantity = int(request.POST.get("quantity", 1))

    if quantity <= 0:
        messages.error(request, "Quantity must be positive.")
        return redirect(request.META.get("HTTP_REFERER", "home"))  # Redirect back

    selection = get_booking_selection(request.session)
    item_name = ""

    if item_id:
        try:
            item = get_object_or_404(RentalItem, pk=item_id, status=ItemStatus.AVAILABLE)
            item_pk_str = str(item.pk)
            item_name = str(item.item_type)

            # Check if adding this quantity exceeds available stock (simple check)
            # Note: Real availability check happens at finalize_booking
            current_quantity = selection["items"].get(item_pk_str, {}).get("quantity", 0)
            # A more robust check would query the DB for available count, but can be slow here.
            # Let's assume we allow adding for now and check properly at finalization.

            selection["items"][item_pk_str] = {
                "quantity": current_quantity + quantity,
                "type": "item",
            }
            messages.success(request, f"Added {quantity} x '{item_name}' to your selection.")

        except Http404:
            messages.error(request, "Item not found or is unavailable.")
            return redirect(request.META.get("HTTP_REFERER", "home"))
        except ValueError:
            messages.error(request, "Invalid item ID.")
            return redirect(request.META.get("HTTP_REFERER", "home"))

    elif set_id:
        try:
            item_set = get_object_or_404(ItemSet, pk=set_id)
            set_pk_str = str(item_set.pk)
            item_name = item_set.name

            # Check available quantity for the set
            # available_set_quantity = item_set.available_quantity
            # current_quantity = selection["sets"].get(set_pk_str, {}).get("quantity", 0)

            # if current_quantity + quantity > available_set_quantity:
            #     messages.warning(
            #         request,
            #         f"Cannot add {quantity} x '{item_name}'. Only {available_set_quantity - current_quantity} more available.",
            #     )
            #     # Adjust quantity to max available if possible, or just prevent adding
            #     # For simplicity, let's just prevent adding more than available for sets here.
            #     # If you want to auto-adjust: quantity = max(0, available_set_quantity - current_quantity)
            #     # if quantity == 0: return redirect(...) # etc.
            #     return redirect(request.META.get("HTTP_REFERER", "home"))

            selection["sets"][set_pk_str] = {
                # "quantity": current_quantity + quantity,
                "quantity": quantity,
                "type": "set",
            }
            messages.success(request, f"Added {quantity} x '{item_name}' set to your selection.")

        except Http404:
            messages.error(request, "Item set not found.")
            return redirect(request.META.get("HTTP_REFERER", "home"))
        except ValueError:
            messages.error(request, "Invalid set ID.")
            return redirect(request.META.get("HTTP_REFERER", "home"))

    else:
        messages.error(request, "No item or set specified.")
        return redirect(request.META.get("HTTP_REFERER", "home"))

    request.session["booking_selection"] = selection
    request.session.modified = True  # Ensure session is saved

    # Redirect to the selection page or back where they came from
    return redirect(reverse("view_booking_selection"))


@login_required
def view_booking_selection(request):
    selection = get_booking_selection(request.session)
    items_in_selection = []
    sets_in_selection = []
    total_price_per_day = decimal.Decimal("0.00")
    total_deposit_estimate = decimal.Decimal("0.00")  # Placeholder for deposit calculation

    # Get item details
    item_ids = [int(pk) for pk in selection["items"].keys()]
    if item_ids:
        items = RentalItem.objects.filter(pk__in=item_ids).select_related("item_type")
        item_map = {item.pk: item for item in items}
        for pk_str, data in selection["items"].items():
            pk = int(pk_str)
            if pk in item_map:
                item = item_map[pk]
                quantity = data["quantity"]
                price = item.item_type.rental_price_per_day * quantity
                items_in_selection.append(
                    {"item": item, "quantity": quantity, "price_per_day": price}
                )
                total_price_per_day += price
                # Add item replacement cost to deposit estimate (example logic)
                total_deposit_estimate += (
                    item.item_type.replacement_cost * quantity * decimal.Decimal("0.3")
                )  # e.g., 30% of replacement

    # Get set details
    set_ids = [int(pk) for pk in selection["sets"].keys()]
    if set_ids:
        item_sets = ItemSet.objects.filter(pk__in=set_ids)
        set_map = {s.pk: s for s in item_sets}
        for pk_str, data in selection["sets"].items():
            pk = int(pk_str)
            if pk in set_map:
                item_set = set_map[pk]
                quantity = data["quantity"]
                price = item_set.base_price * quantity
                sets_in_selection.append(
                    {"set": item_set, "quantity": quantity, "price_per_day": price}
                )
                total_price_per_day += price
                # Add set deposit to estimate
                total_deposit_estimate += (
                    item_set.replacement_deposit * quantity * decimal.Decimal("0.3")
                )  # e.g., 30% of replacement

    # Use the existing BookingForm for dates and payment method
    # The form action will point to 'finalize_booking'
    if request.method == "POST":
        # This view doesn't handle the final booking POST, finalize_booking does.
        # But you might handle quantity updates here if not using separate views/AJAX.
        form = BookingForm()
        pass  # Or handle quantity updates if form submitted here
    else:
        form = BookingForm()  # For dates/payment method

    context = {
        "items_in_selection": items_in_selection,
        "sets_in_selection": sets_in_selection,
        "total_price_per_day": total_price_per_day,
        "total_deposit_estimate": total_deposit_estimate,  # Pass estimate to template
        "form": form,  # Date/Payment form
    }
    return render(request, "core/clients/pages/booking_selection.html", context)


@require_POST
@login_required
def remove_from_selection(request, item_type, pk):
    selection = get_booking_selection(request.session)
    pk_str = str(pk)
    removed = False

    if item_type == "item" and pk_str in selection["items"]:
        del selection["items"][pk_str]
        removed = True
        messages.success(request, "Item removed from selection.")
    elif item_type == "set" and pk_str in selection["sets"]:
        del selection["sets"][pk_str]
        removed = True
        messages.success(request, "Set removed from selection.")
    else:
        messages.error(request, "Item or set not found in selection.")

    if removed:
        request.session["booking_selection"] = selection
        request.session.modified = True

    return redirect(reverse("view_booking_selection"))


@require_POST
@login_required
def update_selection_quantity(request, item_type, pk):
    selection = get_booking_selection(request.session)
    pk_str = str(pk)
    new_quantity = int(request.POST.get("quantity", 0))

    if new_quantity <= 0:
        # Treat as removal if quantity is zero or less
        return remove_from_selection(request, item_type, pk)

    updated = False
    if item_type == "item" and pk_str in selection["items"]:
        # Add check against available stock if needed here (complex without dates)
        selection["items"][pk_str]["quantity"] = new_quantity
        updated = True
        messages.success(request, "Item quantity updated.")
    elif item_type == "set" and pk_str in selection["sets"]:
        try:
            item_set = ItemSet.objects.get(pk=pk)
            # if new_quantity > item_set.available_quantity:
            #     messages.error(
            #         request,
            #         f"Only {item_set.available_quantity} of '{item_set.name}' set(s) available.",
            #     )
            # else:
            selection["sets"][pk_str]["quantity"] = new_quantity
            updated = True
            messages.success(request, "Set quantity updated.")
        except ItemSet.DoesNotExist:
            messages.error(request, "Set not found.")

    else:
        messages.error(request, "Item or set not found in selection.")

    if updated:
        request.session["booking_selection"] = selection
        request.session.modified = True

    return redirect(reverse("view_booking_selection"))


@require_POST  # Final booking should be a POST request
@login_required
def finalize_booking(request):
    selection = get_booking_selection(request.session)
    form = BookingForm(request.POST, request.FILES)

    if not selection["items"] and not selection["sets"]:
        messages.error(request, "Your booking selection is empty.")
        return redirect(reverse("view_booking_selection"))

    if form.is_valid():
        start_date = form.cleaned_data["start_date"]
        end_date = form.cleaned_data["end_date"]
        payment_method = form.cleaned_data["payment_method"]
        payment_slip_file = form.cleaned_data.get("payment_slip")  # Use .get() as it might be None

        # --- Data Validation & Calculation ---
        duration_days = max(1, (end_date - start_date).days + 1)
        total_rental_cost = decimal.Decimal("0.00")
        total_deposit = decimal.Decimal("0.00")
        items_to_book_details = []  # Store details for RentalItemDetail creation
        sets_to_book_details = []  # Store details for RentalSetDetail creation
        customer = get_object_or_404(Customer, user=request.user)

        # --- Transaction for atomicity ---
        try:
            with transaction.atomic():
                # == Availability Check and Cost Calculation ==

                # 1. Check Item Sets First (as they consume individual items)
                set_ids = [int(pk) for pk in selection["sets"].keys()]
                required_item_counts = defaultdict(int)  # Track total required count per item type

                if set_ids:
                    item_sets = ItemSet.objects.filter(pk__in=set_ids).prefetch_related(
                        "itemsetcomponent_set__item_type"
                    )
                    set_map = {s.pk: s for s in item_sets}
                    for pk_str, data in selection["sets"].items():
                        pk = int(pk_str)
                        if pk not in set_map:
                            raise ValueError(
                                f"Selected Item Set (ID: {pk}) not found."
                            )  # Abort transaction

                        item_set = set_map[pk]
                        quantity_needed = data["quantity"]

                        # Check overall set availability for the dates (complex - requires checking components)
                        # Simplified check: Assume ItemSet.available_quantity was checked earlier.
                        # A full check iterates components and checks RentalItem availability.
                        # For now, rely on the earlier check in add/update.

                        # Calculate cost and deposit for the set
                        set_price_per_day = item_set.base_price
                        total_rental_cost += set_price_per_day * quantity_needed
                        total_deposit += item_set.replacement_deposit * quantity_needed

                        sets_to_book_details.append(
                            {
                                "item_set": item_set,
                                "quantity": quantity_needed,
                                "rented_price_per_day": set_price_per_day,
                            }
                        )

                        # Aggregate required item types for conflict checking later
                        for component in item_set.itemsetcomponent_set.all():
                            required_item_counts[component.item_type.pk] += (
                                component.quantity * quantity_needed
                            )

                # 2. Check Individual Items
                item_ids = [int(pk) for pk in selection["items"].keys()]
                if item_ids:
                    items = RentalItem.objects.filter(pk__in=item_ids).select_related("item_type")
                    item_map = {item.pk: item for item in items}
                    for pk_str, data in selection["items"].items():
                        pk = int(pk_str)
                        if pk not in item_map:
                            raise ValueError(f"Selected Item (ID: {pk}) not found.")  # Abort

                        item = item_map[pk]
                        quantity_needed = data["quantity"]  # Should always be 1 for specific items

                        # if quantity_needed > 1:
                        #     # This flow assumes adding specific items, so quantity should be 1
                        #     # If you allow adding multiple of the *same* specific item, logic needs change
                        #     raise ValueError(
                        #         f"Invalid quantity ({quantity_needed}) for specific item {item.pk}."
                        #     )

                        # Check availability for THIS specific item
                        if item.status != ItemStatus.AVAILABLE:
                            raise ValueError(f"Item '{item}' is no longer available.")

                        overlapping = RentalItemDetail.objects.filter(
                            item=item,
                            rental__end_date__gte=start_date,
                            rental__start_date__lte=end_date,
                            rental__payment_status__in=[
                                PaymentStatus.PAID,
                                PaymentStatus.PENDING,
                                PaymentStatus.PARTIAL,
                            ],  # Check active/pending rentals
                        ).exists()

                        if overlapping:
                            raise ValueError(
                                f"Item '{item}' is already booked for the selected dates."
                            )

                        # Calculate cost and deposit for the item
                        item_price_per_day = item.item_type.rental_price_per_day
                        total_rental_cost += item_price_per_day * quantity_needed
                        # Example deposit logic for individual items (e.g., 30% of replacement)
                        total_deposit += (
                            item.item_type.replacement_cost
                            * quantity_needed
                            * decimal.Decimal("0.30")
                        )

                        items_to_book_details.append(
                            {
                                "item": item,
                                "quantity": quantity_needed,
                                "rented_price_per_day": item_price_per_day,
                            }
                        )

                        # Add to required item counts if needed for complex availability checks
                        # required_item_counts[item.item_type.pk] += 1

                # --- Create Database Records ---

                # a. Create RentalTransaction
                # Note: total_rental_cost here is per day. Multiply by duration for full cost.
                # The deposit calculation might need refinement (e.g., min deposit)
                final_total_deposit = max(
                    total_deposit, decimal.Decimal("20.00")
                )  # Example minimum deposit

                rental_transaction = RentalTransaction.objects.create(
                    customer=customer,
                    start_date=start_date,
                    end_date=end_date,
                    total_deposit=final_total_deposit,
                    payment_status=PaymentStatus.PENDING,
                )

                # b. Create RentalSetDetail records (if any sets were selected)
                created_set_details = {}  # To link items to their set rental
                for set_detail_data in sets_to_book_details:
                    set_detail = RentalSetDetail.objects.create(
                        rental=rental_transaction,
                        item_set=set_detail_data["item_set"],
                        quantity=set_detail_data["quantity"],
                        rented_price_per_day=set_detail_data["rented_price_per_day"],
                    )
                    created_set_details[set_detail_data["item_set"].pk] = set_detail

                    # --- Allocate specific items for the set components ---
                    # This is the complex part: finding available items of the required types.
                    item_set = set_detail_data["item_set"]
                    set_quantity = set_detail_data["quantity"]
                    for component in item_set.itemsetcomponent_set.all():
                        items_needed_for_component = component.quantity * set_quantity
                        # Find available items of this type NOT already booked for these dates
                        available_items = RentalItem.objects.filter(
                            item_type=component.item_type, status=ItemStatus.AVAILABLE
                        ).exclude(  # Exclude items booked in the date range
                            rentalitemdetail__rental__end_date__gte=start_date,
                            rentalitemdetail__rental__start_date__lte=end_date,
                            rentalitemdetail__rental__payment_status__in=[
                                PaymentStatus.PAID,
                                PaymentStatus.PENDING,
                                PaymentStatus.PARTIAL,
                            ],
                        )[
                            :items_needed_for_component
                        ]  # Limit to the number needed

                        # if len(available_items) < items_needed_for_component:
                        #     raise ValueError(
                        #         f"Not enough available '{component.item_type}' items for the '{item_set.name}' set for the selected dates."
                        #     )

                        # Create RentalItemDetail for each allocated item and link to set detail
                        for allocated_item in available_items:
                            RentalItemDetail.objects.create(
                                rental=rental_transaction,
                                item=allocated_item,
                                quantity=quantity_needed,  # Each detail is for one specific item instance
                                rented_price_per_day=component.item_type.rental_price_per_day,  # Or use a set-component specific price?
                                set_rental=set_detail,  # Link back to the set detail
                            )

                # c. Create RentalItemDetail records for individually selected items
                for item_detail_data in items_to_book_details:
                    item = item_detail_data["item"]
                    RentalItemDetail.objects.create(
                        rental=rental_transaction,
                        item=item,
                        quantity=item_detail_data["quantity"],
                        rented_price_per_day=item_detail_data["rented_price_per_day"],
                        set_rental=None,  # Not part of a set rental
                    )

                # d. Create Initial Payment Record (for the deposit)
                Payment.objects.create(
                    rental=rental_transaction,
                    amount=final_total_deposit,
                    payment_method=payment_method,
                    payment_type=PaymentType.DEPOSIT,
                    payment_slip=payment_slip_file,
                    notes=f"Initial deposit for Rental #{rental_transaction.id}",
                )

                # --- Success ---
                clear_booking_selection(request.session)  # Clear the cart/selection
                messages.success(
                    request,
                    "Booking created successfully! Please complete payment if required.",
                )
                return redirect(
                    reverse("booking_success", kwargs={"transaction_pk": rental_transaction.pk})
                )

        except ValueError as ve:
            # Handle specific validation errors (like availability)
            messages.error(request, f"Booking failed: {ve}")
            # Re-render selection page with form errors
            context = get_selection_context_for_render(request, form)  # Helper to get context
            return render(request, "core/clients/pages/booking_selection.html", context)

        except Exception as e:
            # Handle unexpected errors during transaction
            messages.error(request, f"An unexpected error occurred while creating the booking: {e}")
            context = get_selection_context_for_render(request, form)  # Helper to get context
            return render(request, "core/clients/pages/booking_selection.html", context)

    else:
        # Form is invalid (e.g., dates are wrong)
        messages.error(request, "Please correct the errors in the form below.")
        context = get_selection_context_for_render(request, form)  # Helper to get context
        return render(request, "core/clients/pages/booking_selection.html", context)


def get_selection_context_for_render(request, form):
    """Helper to rebuild context for rendering the selection page after errors."""
    selection = get_booking_selection(request.session)
    items_in_selection = []
    sets_in_selection = []
    total_price_per_day = decimal.Decimal("0.00")
    total_deposit_estimate = decimal.Decimal("0.00")

    item_ids = [int(pk) for pk in selection["items"].keys()]
    if item_ids:
        items = RentalItem.objects.filter(pk__in=item_ids).select_related("item_type")
        item_map = {item.pk: item for item in items}
        for pk_str, data in selection["items"].items():
            pk = int(pk_str)
            if pk in item_map:
                item = item_map[pk]
                quantity = data["quantity"]
                price = item.item_type.rental_price_per_day * quantity
                items_in_selection.append(
                    {"item": item, "quantity": quantity, "price_per_day": price}
                )
                total_price_per_day += price
                total_deposit_estimate += (
                    item.item_type.replacement_cost * quantity * decimal.Decimal("0.3")
                )

    set_ids = [int(pk) for pk in selection["sets"].keys()]
    if set_ids:
        item_sets = ItemSet.objects.filter(pk__in=set_ids)
        set_map = {s.pk: s for s in item_sets}
        for pk_str, data in selection["sets"].items():
            pk = int(pk_str)
            if pk in set_map:
                item_set = set_map[pk]
                quantity = data["quantity"]
                price = item_set.base_price * quantity
                sets_in_selection.append(
                    {"set": item_set, "quantity": quantity, "price_per_day": price}
                )
                total_price_per_day += price
                total_deposit_estimate += (
                    item_set.replacement_deposit * quantity * decimal.Decimal("0.3")
                )

    return {
        "items_in_selection": items_in_selection,
        "sets_in_selection": sets_in_selection,
        "total_price_per_day": total_price_per_day,
        "total_deposit_estimate": total_deposit_estimate,
        "form": form,
    }


@login_required
def booking_success_view(request, transaction_pk):
    # Retrieve the transaction and related details
    rental_transaction = get_object_or_404(
        RentalTransaction.objects.prefetch_related(
            "rentalitemdetail_set__item__item_type",  # Get item details
            "rentalsetdetail_set__item_set",  # Get set details
        ),
        pk=transaction_pk,
        customer__user=request.user,
    )

    # Separate details for individual items vs items part of sets
    individual_item_details = rental_transaction.rentalitemdetail_set.filter(
        set_rental__isnull=True
    )
    set_details = (
        rental_transaction.rentalsetdetail_set.all()
    )  # Includes items via related_name if needed

    context = {
        "transaction": rental_transaction,
        "individual_item_details": individual_item_details,
        "set_details": set_details,
        "PaymentStatus": PaymentStatus,  # Pass Enum for template comparisons
    }
    return render(request, "core/clients/pages/booking_success.html", context)


@login_required
def dashboard(request):
    """Displays the main admin dashboard with summary statistics."""
    today = timezone.now().date()
    current_month_start = today.replace(day=1)

    # Summary Counts
    pending_approvals_count = RentalTransaction.objects.filter(
        payment_status=PaymentStatus.PENDING
    ).count()
    active_rentals_count = RentalTransaction.objects.filter(
        payment_status=PaymentStatus.PAID, start_date__lte=today, end_date__gte=today
    ).count()
    total_customers_count = Customer.objects.count()
    total_items_count = RentalItem.objects.count()
    available_items_count = RentalItem.objects.filter(status=ItemStatus.AVAILABLE).count()
    total_sets_count = ItemSet.objects.count()

    available_sets_count = 0
    # This loop can be inefficient for many sets. Consider optimizing if performance is an issue.
    for item_set_obj in ItemSet.objects.all():
        if item_set_obj.available_quantity > 0:
            available_sets_count += 1

    # New Data for Dashboard:
    # 1. Upcoming Returns (e.g., in the next 7 days)
    upcoming_return_date_limit = today + timedelta(days=7)
    upcoming_returns = (
        RentalTransaction.objects.filter(
            end_date__gte=today,
            end_date__lte=upcoming_return_date_limit,
            payment_status__in=[PaymentStatus.PAID, PaymentStatus.PARTIAL],  # Active rentals
        )
        .select_related("customer__user")
        .order_by("end_date")[:5]
    )  # Limit to 5 for display

    # 2. Overdue Rentals
    overdue_rentals = (
        RentalTransaction.objects.filter(
            end_date__lt=today,
            payment_status__in=[
                PaymentStatus.PAID,
                PaymentStatus.PARTIAL,
            ],  # Was active, now overdue
        )
        .select_related("customer__user")
        .order_by("end_date")[:5]
    )  # Limit to 5

    # 3. Recent Bookings
    recent_bookings = RentalTransaction.objects.select_related("customer__user").order_by(
        "-created_at"
    )[:5]

    # A more accurate way to get revenue would be from actual payments made for completed/ongoing rentals.
    # For simplicity, let's sum `amount_paid` for transactions created this month.
    # This is still an approximation as `amount_paid` could include deposits.
    revenue_this_month_approx = Payment.objects.filter(
        rental__created_at__gte=current_month_start,
        # rental__payment_status__in=[PaymentStatus.PAID, PaymentStatus.PARTIAL] # Consider status
    ).aggregate(total=Sum("amount"))["total"] or decimal.Decimal("0.00")

    total_deposits_held = RentalTransaction.objects.filter(
        payment_status__in=[PaymentStatus.PAID, PaymentStatus.PARTIAL],
        end_date__gte=today,  # Active rentals
    ).aggregate(total_deposits=Sum("total_deposit"))["total_deposits"] or decimal.Decimal("0.00")

    context = {
        "pending_approvals_count": pending_approvals_count,
        "active_rentals_count": active_rentals_count,
        "total_customers_count": total_customers_count,
        "total_items_count": total_items_count,
        "available_items_count": available_items_count,
        "total_sets_count": total_sets_count,
        "available_sets_count": available_sets_count,
        # New context variables
        "upcoming_returns": upcoming_returns,
        "overdue_rentals": overdue_rentals,
        "recent_bookings": recent_bookings,
        "revenue_this_month_approx": revenue_this_month_approx,
        "total_deposits_held": total_deposits_held,
    }
    return render(request, "core/dashboard/pages/dashboard.html", context)


def manage_profile(request):
    return render(request, "core/dashboard/pages/manage-profile.html")


def manage_users(request):
    query = request.GET.get("search", "")
    users = Account.objects.all().order_by("-date_joined")

    if query:
        # Search across multiple fields
        users = users.filter(Q(email__icontains=query))

    msg = "Are you sure you want to delete this user?"
    context = {
        "users": users,
        "delete_confirm_msg": msg,
    }
    return render(request, "core/dashboard/pages/manage-users.html", context)


def delete_user(request, pk):
    user = Account.objects.get(pk=pk)

    if request.method == "POST":
        user.delete()
        messages.success(request, f"User '{user.email}' deleted successfully.")
        return redirect("manage-users")

    return render(request, "core/dashboard/pages/manage-users.html")


def manage_employees(request):
    # employee = Employee.objects.get(user=request.user)
    query = request.GET.get("search", "")
    employees = Employee.objects.all().order_by("-updated_at")

    if query:
        # Search across multiple fields
        employees = employees.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(phone_number__icontains=query)
        )

    if request.method == "POST":
        form = EmployeeForm(request.POST)
        email = request.POST.get("email")
        password = request.POST.get("password")

        if form.is_valid():
            hashed_password = make_password(password)
            user = Account.objects.create(email=email, password=hashed_password)
            employee = form.save(commit=True)
            employee.user = user
            employee.save()
            user.save()
            messages.success(request, "Employee created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        # "employee": employee,
        "employees": employees,
        "delete_confirm_msg": "Are you sure you want to delete this employee?",
    }
    return render(request, "core/dashboard/pages/manage-employees.html", context)


def edit_employee(request, pk):
    # Get the employee to update
    employee = Employee.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    emp_form = EmployeeForm(data, instance=employee)  # Bind the data to the existing instance

    if request.method == "POST":
        if emp_form.is_valid():
            emp_form.save()
            messages.success(request, "Employee updated successfully.")
        else:
            return JsonResponse({"error": emp_form.errors}, status=400)
        return redirect("manage-employees")
    else:
        emp_form = EmployeeForm(data, instance=employee)

    return render(request, "core/dashboard/pages/manage-employees.html")


def delete_employee(request, pk):
    employee = Employee.objects.get(pk=pk)
    user = None

    if employee.user:
        user = Account.objects.get(id=employee.user.id)

    if request.method == "POST":
        employee.delete()

        messages.success(
            request, f"Employee '{employee.first_name} {employee.last_name}' deleted successfully."
        )

        if user is not None:
            user.delete()
        return redirect("manage-employees")

    return render(request, "core/dashboard/pages/manage-employees.html")


def manage_customers(request):
    query = request.GET.get("search", "")
    customers = Customer.objects.all().order_by("-updated_at")

    if query:
        # Search across multiple fields
        customers = customers.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(phone__icontains=query)
        )

    if request.method == "POST":
        form = CustomerForm(request.POST)
        email = request.POST.get("email")
        password = request.POST.get("password")

        if form.is_valid():
            hashed_password = make_password(password)
            user = Account.objects.create(email=email, password=hashed_password)
            customer = Customer(
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                phone_number=form.cleaned_data["phone_number"],
                address=form.cleaned_data.get("address", ""),
                user=user,
            )
            customer.save()

            messages.success(request, "Customer created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "customers": customers,
        "delete_confirm_msg": "Are you sure you want to delete this customer?",
    }
    return render(request, "core/dashboard/pages/manage-customers.html", context)


def edit_customer(request, pk):
    customer = Customer.objects.get(id=pk)
    data = QueryDict(request.body)
    customer_form = CustomerForm(data, instance=customer)

    if request.method == "POST":
        if customer_form.is_valid():
            customer_form.save()
            messages.success(request, "Customer updated successfully.")
        else:
            return JsonResponse({"error": customer_form.errors}, status=400)
        return redirect("manage-customers")
    else:
        customer_form = CustomerForm(data, instance=customer)

    return render(request, "core/dashboard/pages/manage-customers.html")


def delete_customer(request, pk):
    customer = Customer.objects.get(pk=pk)

    if request.method == "POST":
        customer.delete()
        messages.success(
            request, f"Customer '{customer.first_name} {customer.last_name}' deleted successfully."
        )
        return redirect("manage-customers")

    return render(request, "core/dashboard/pages/manage-customers.html")


def manage_rental_item_types(request):
    query = request.GET.get("search", "")
    rental_item_types = RentalItemType.objects.all().order_by("-updated_at")
    rental_item_type_choices = ItemType.choices

    if query:
        # Search across multiple fields
        rental_item_types = RentalItemType.filter(Q(type_name__icontain=query))

    if request.method == "POST":
        form = RentalItemTypeForm(request.POST)

        if form.is_valid():
            rental_item_type = form.save(commit=True)
            rental_item_type.save()
            messages.success(request, "Rental item type created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "rental_item_types": rental_item_types,
        "rental_item_type_choices": rental_item_type_choices,
        "delete_confirm_msg": "Are you sure you want to delete this rental item type?",
    }
    return render(request, "core/dashboard/pages/manage-rental-item-types.html", context)


def edit_rental_item_type(request, pk):
    rental_item_type = RentalItemType.objects.get(id=pk)
    data = QueryDict(request.body)
    rental_item_type_form = RentalItemTypeForm(data, instance=rental_item_type)

    if request.method == "POST":
        if rental_item_type_form.is_valid():
            rental_item_type_form.save()
            messages.success(request, "Rental item type updated successfully.")
        else:
            return JsonResponse({"error": rental_item_type_form.errors}, status=400)
        return redirect("manage-rental-item-types")

    rental_item_type_form = RentalItemTypeForm(data, instance=rental_item_type)

    return render(request, "core/dashboard/pages/manage-rental-item-types.html")


def delete_rental_item_type(request, pk):
    rental_item_type = RentalItemType.objects.get(pk=pk)

    if request.method == "POST":
        rental_item_type.delete()
        messages.success(
            request, f"Rental Item Type '{rental_item_type.type_name}' deleted successfully."
        )
        return redirect("manage-rental-item-types")

    return render(request, "core/dashboard/pages/manage-rental-item-types.html")


def manage_rental_items(request):
    query = request.GET.get("search", "")
    rental_items = RentalItem.objects.all().order_by("-updated_at")
    rental_item_types = RentalItemType.objects.all().order_by("-updated_at")
    rental_item_status_choices = ItemStatus.choices

    prefix = "RI"
    date_str = timezone.now().strftime("%Y%m%d")
    sequence = str(uuid.uuid4().int)[-3:]

    if query:
        # Search across multiple fields
        rental_items = rental_items.filter(Q(serial_number__icontains=query))

    if request.method == "POST":
        form = RentalItemForm(request.POST, request.FILES)

        if form.is_valid():
            rental_item = form.save(commit=False)
            rental_item.serial_number = f"{prefix}-{date_str}-{sequence}"
            rental_item.save()
            messages.success(request, "Rental item created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "rental_items": rental_items,
        "rental_item_types": rental_item_types,
        "rental_item_status_choices": rental_item_status_choices,
        "delete_confirm_msg": "Are you sure you want to delete this rental item?",
    }
    return render(request, "core/dashboard/pages/manage-rental-items.html", context)


def edit_rental_item(request, pk):
    rental_item = RentalItem.objects.get(id=pk)

    if request.method == "POST":
        form = RentalItemForm(request.POST, request.FILES, instance=rental_item)

        if form.is_valid():
            form.save()
            messages.success(request, "Rental item updated successfully.")
        else:
            return JsonResponse({"error": form.errors}, status=400)
        return redirect("manage-rental-items")
    else:
        form = RentalItemForm(instance=rental_item)

    return render(request, "core/dashboard/pages/manage-rental-items.html")


def delete_rental_item(request, pk):
    rental_item = RentalItem.objects.get(pk=pk)

    if request.method == "POST":
        rental_item.delete()
        messages.success(
            request, f"Rental Item '{rental_item.serial_number}' deleted successfully."
        )
        return redirect("manage-rental-items")

    return render(request, "core/dashboard/pages/manage-rental-items.html")


def manage_accessories(request):
    query = request.GET.get("search", "")
    accessories = Accessory.objects.all().order_by("-updated_at")
    rental_item_types = RentalItemType.objects.all().order_by("-updated_at")
    rental_item_status_choices = ItemStatus.choices

    if query:
        # Search across multiple fields
        accessories = accessories.filter(Q(accessory_name__icontains=query))

    if request.method == "POST":
        form = AccessoryForm(request.POST)

        if form.is_valid():
            accessory = form.save(commit=False)
            accessory.save()
            messages.success(request, "Accessory item created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "accessories": accessories,
        "rental_item_types": rental_item_types,
        "rental_item_status_choices": rental_item_status_choices,
        "delete_confirm_msg": "Are you sure you want to delete this accessory?",
    }
    return render(request, "core/dashboard/pages/manage-accessories.html", context)


def edit_accessory(request, pk):
    accessory = Accessory.objects.get(id=pk)
    data = QueryDict(request.body)
    accessory_form = AccessoryForm(data, instance=accessory)

    if request.method == "POST":
        if accessory_form.is_valid():
            accessory_form.save()
            messages.success(request, "Accessory item updated successfully.")
        else:
            return JsonResponse({"error": accessory_form.errors}, status=400)
        return redirect("manage-accessories")
    else:
        accessory_form = AccessoryForm(data, instance=accessory)

    return render(request, "core/dashboard/pages/manage-accessories.html")


def delete_accessory(request, pk):
    accessory = Accessory.objects.get(pk=pk)

    if request.method == "POST":
        accessory.delete()
        messages.success(request, f"Accessory '{accessory.accessory_name}' deleted successfully.")
        return redirect("manage-accessories")

    return render(request, "core/dashboard/pages/manage-accessories.html")


@login_required
def manage_item_sets(request):
    """
    Displays a list of all Item Sets and handles search.
    """
    query = request.GET.get("search", "")
    item_sets = ItemSet.objects.all().order_by("name")

    if query:
        item_sets = item_sets.filter(Q(name__icontains=query) | Q(description__icontains=query))

    if request.method == "POST":
        form = ItemSetForm(request.POST, request.FILES)
        if form.is_valid():
            item_set = form.save()
            messages.success(request, "Item set created successfully! Now add components.")
            return redirect("add_components", set_id=item_set.id)
    else:
        form = ItemSetForm()

    context = {
        "item_sets": item_sets,
        "search_query": query,
        "delete_confirm_msg": "Are you sure you want to delete this item set and all its components?",
    }
    return render(request, "core/dashboard/pages/manage-item-sets.html", context)


def edit_item_set(request, pk):
    item_set = ItemSet.objects.get(id=pk)
    item_set_form = ItemSetForm(request.POST, request.FILES, instance=item_set)

    if request.method == "POST":
        if item_set_form.is_valid():
            item_set_form.save()
            messages.success(request, "Edit item set updated successfully.")
        else:
            return JsonResponse({"error": item_set_form.errors}, status=400)
        return redirect("manage-item-sets")
    else:
        item_set_form = ItemSetForm(instance=item_set)

    return render(request, "core/dashboard/pages/manage-item-sets.html")


@login_required
def add_components(request, set_id):
    item_set = get_object_or_404(ItemSet, pk=set_id)

    if request.method == "POST":
        form = ItemSetComponentForm(request.POST)
        if form.is_valid():
            component = form.save(commit=False)
            component.item_set = item_set
            component.save()
            messages.success(request, "Component added successfully!")
            return redirect("add_components", set_id=set_id)
    else:
        form = ItemSetComponentForm(initial={"item_set": item_set})

    # Get existing components
    components = ItemSetComponent.objects.filter(item_set=item_set)

    context = {
        "form": form,
        "item_set": item_set,
        "components": components,
        "item_types": RentalItemType.objects.all(),
        "title": f"Add Components to {item_set.name}",
    }
    return render(request, "core/dashboard/modals/add/add-item-set-components.html", context)


@login_required
def delete_item_set(request, pk):
    item_set = get_object_or_404(ItemSet, pk=pk)
    if request.method == "POST":
        try:
            item_set.delete()
            messages.success(request, "Item Set deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting item set: {e}")
        return redirect("manage-item-sets")
    return redirect("manage-item-sets")


@login_required
def delete_item_set_component(request, pk):
    item_set_component = get_object_or_404(ItemSetComponent, pk=pk)
    item_set_id = item_set_component.item_set.id

    if request.method == "POST":
        try:
            item_set_component.delete()
            messages.success(request, "Item Set Component deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting item set component: {e}")
        return redirect("add_components", set_id=item_set_id)
    return redirect("add_components", set_id=item_set_id)


@login_required
def manage_rental_approvals(request):
    """Lists rental transactions pending approval (e.g., pending payment confirmation)."""

    # Fetch transactions that are PENDING and potentially have a payment slip
    # Prefetch related objects for efficiency in the template
    pending_transactions = (
        RentalTransaction.objects.filter(payment_status=PaymentStatus.PENDING)
        .select_related("customer__user")  # Get customer details efficiently
        .prefetch_related(
            "rentalitemdetail_set__item__item_type",  # Get details of individual items rented
            "rentalsetdetail_set__item_set",  # Get details of sets rented
            Prefetch(  # Prefetch the relevant payment (deposit with slip)
                "payment_set",
                queryset=Payment.objects.filter(payment_type=PaymentType.DEPOSIT).order_by(
                    "-created_at"
                ),
                to_attr="deposit_payments",  # Assign prefetched payments to this attribute
            ),
        )
        .order_by("start_date")
    )  # Order by start date, oldest first

    context = {
        "pending_transactions": pending_transactions,
        "title": "Rental Approvals",  # For dashboard page title
    }
    return render(request, "core/dashboard/pages/manage_rental_approvals.html", context)


@require_POST
@login_required
def approve_rental(request, transaction_pk):
    """Approves a rental transaction, typically by updating its payment status."""
    rental_transaction = get_object_or_404(RentalTransaction, pk=transaction_pk)

    # Check if it's actually pending
    if rental_transaction.payment_status != PaymentStatus.PENDING:
        messages.warning(request, f"Rental #{rental_transaction.id} is not pending approval.")
        return redirect("manage_rental_approvals")

    try:
        # --- Approval Logic ---
        # For now, simply update the status to PAID.
        # You might add more complex logic here later (e.g., verifying payment details).
        rental_transaction.payment_status = PaymentStatus.PAID
        rental_transaction.save()

        # Optional: Update the status of the associated deposit Payment record if needed
        # deposit_payment = rental_transaction.payment_set.filter(payment_type=PaymentType.DEPOSIT).first()
        # if deposit_payment:
        #     # Update payment status if your Payment model has one
        #     pass

        messages.success(request, f"Rental #{rental_transaction.id} approved successfully.")

    except Exception as e:
        messages.error(
            request, f"An error occurred while approving rental #{rental_transaction.id}: {e}"
        )

    return redirect("manage_rental_approvals")


@require_POST
@login_required
def reject_rental(request, transaction_pk):
    """Rejects a rental transaction by updating its payment status to Cancelled."""
    rental_transaction = get_object_or_404(RentalTransaction, pk=transaction_pk)

    # Check if it's actually pending
    if rental_transaction.payment_status != PaymentStatus.PENDING:
        messages.warning(request, f"Rental #{rental_transaction.id} is not pending rejection.")
        return redirect("manage_rental_approvals")

    try:
        # --- Rejection Logic ---
        # Update the status to CANCELLED.
        # You might want to add a note or reason field to the transaction later.
        rental_transaction.payment_status = PaymentStatus.CANCELLED
        rental_transaction.save()

        # Optional: Add logic here to notify the customer about the rejection.

        messages.success(request, f"Rental #{rental_transaction.id} rejected successfully.")

    except Exception as e:
        messages.error(
            request, f"An error occurred while rejecting rental #{rental_transaction.id}: {e}"
        )

    return redirect("manage_rental_approvals")


@login_required
def manage_rental_transactions(request):
    """Lists all rental transactions with filtering."""

    transactions_list = (
        RentalTransaction.objects.select_related("customer__user")
        .prefetch_related(
            "rentalitemdetail_set__item__item_type",
            "rentalsetdetail_set__item_set",
            "payment_set",  # Prefetch payments for potential display
        )
        .order_by("-start_date", "-created_at")
    )  # Default order

    # --- Filtering Logic ---
    status_filter = request.GET.get("status")
    customer_filter = request.GET.get("customer")  # Search by customer name/email
    date_from_filter = request.GET.get("date_from")
    date_to_filter = request.GET.get("date_to")

    if status_filter:
        transactions_list = transactions_list.filter(payment_status=status_filter)

    if customer_filter:
        transactions_list = transactions_list.filter(
            Q(customer__first_name__icontains=customer_filter)
            | Q(customer__last_name__icontains=customer_filter)
            | Q(customer__user__email__icontains=customer_filter)
        )

    if date_from_filter:
        try:
            # Assuming YYYY-MM-DD format from a date input
            start_date = timezone.datetime.strptime(date_from_filter, "%Y-%m-%d").date()
            transactions_list = transactions_list.filter(start_date__gte=start_date)
        except ValueError:
            messages.warning(request, "Invalid 'Date From' format. Please use YYYY-MM-DD.")

    if date_to_filter:
        try:
            end_date = timezone.datetime.strptime(date_to_filter, "%Y-%m-%d").date()
            # Filter for rentals ENDING on or before the specified date
            transactions_list = transactions_list.filter(end_date__lte=end_date)
        except ValueError:
            messages.warning(request, "Invalid 'Date To' format. Please use YYYY-MM-DD.")

    # --- Pagination ---
    paginator = Paginator(transactions_list, 25)  # Show 25 transactions per page
    page_number = request.GET.get("page")
    try:
        transactions = paginator.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        transactions = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        transactions = paginator.page(paginator.num_pages)

    context = {
        "transactions": transactions,
        "payment_status_choices": PaymentStatus.choices,  # Pass choices for filter dropdown
        "current_status_filter": status_filter,  # Pass current filter values back to template
        "current_customer_filter": customer_filter,
        "current_date_from_filter": date_from_filter,
        "current_date_to_filter": date_to_filter,
        "title": "All Rental Transactions",
    }
    return render(request, "core/dashboard/pages/manage_rental_transactions.html", context)


# --- Return Processing Views ---
@login_required
def manage_returns(request):
    """Lists active rentals that are due or overdue for return and settlement."""
    today = timezone.now().date()
    returnable_transactions = (
        RentalTransaction.objects.filter(
            Q(payment_status=PaymentStatus.PAID) | Q(payment_status=PaymentStatus.PARTIAL),
            # Consider listing all PAID/PARTIAL, or filter by end_date for more specific targeting
            # end_date__lte=today + timedelta(days=7) # Example: items due within next 7 days or past due
        )
        .select_related("customer__user")
        .prefetch_related("rentalitemdetail_set__item", "rentalsetdetail_set__item_set")
        .order_by("end_date", "created_at")
    )

    query = request.GET.get("search", "")
    if query:
        returnable_transactions = returnable_transactions.filter(
            Q(id__icontains=query)
            | Q(customer__first_name__icontains=query)
            | Q(customer__last_name__icontains=query)
            | Q(customer__user__email__icontains=query)
        )

    paginator = Paginator(returnable_transactions, 20)
    page_number = request.GET.get("page")
    try:
        transactions_page = paginator.page(page_number)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)

    context = {
        "transactions": transactions_page,
        "title": "Process Returns & Settlements",
        "search_query": query,
        "PaymentStatus": PaymentStatus,  # Pass Enum for template comparisons
    }
    return render(request, "core/dashboard/pages/manage_returns.html", context)


@login_required
def process_return(request, transaction_pk):
    transaction_obj = get_object_or_404(
        RentalTransaction.objects.select_related("customer__user").prefetch_related(
            "rentalitemdetail_set__item__item_type",
            "rentalsetdetail_set__item_set__itemsetcomponent_set__item_type",
            "payment_set",
        ),
        pk=transaction_pk,
    )

    if transaction_obj.payment_status == PaymentStatus.COMPLETED:
        messages.info(
            request, f"Rental #{transaction_obj.id} has already been completed and settled."
        )
        return redirect("manage_returns")
    if transaction_obj.payment_status == PaymentStatus.CANCELLED:
        messages.warning(
            request,
            f"Rental #{transaction_obj.id} was cancelled and cannot be processed for return.",
        )
        return redirect("manage_returns")
    if transaction_obj.payment_status == PaymentStatus.PENDING:
        messages.warning(
            request, f"Rental #{transaction_obj.id} is still pending initial approval/payment."
        )
        return redirect("manage_rental_approvals")

    initial_balance_due = (
        transaction_obj.total_rental_cost + transaction_obj.total_fines
    ) - transaction_obj.amount_paid
    other_payments_made = transaction_obj.amount_paid - transaction_obj.total_deposit

    if request.method == "POST":
        form = ReturnSettlementForm(request.POST, request.FILES)
        if form.is_valid():
            additional_fine = form.cleaned_data.get(
                "additional_fine_amount", decimal.Decimal("0.00")
            )
            fine_description = form.cleaned_data.get("fine_description")
            final_payment_method = form.cleaned_data.get("final_payment_method")
            final_payment_slip = form.cleaned_data.get("final_payment_slip")
            notes_on_return = form.cleaned_data.get("notes_on_return")

            try:
                with transaction.atomic():
                    if additional_fine and additional_fine > 0:
                        transaction_obj.total_fines += additional_fine

                    transaction_obj.save()

                    final_amount_to_settle = (
                        transaction_obj.total_rental_cost + transaction_obj.total_fines
                    ) - transaction_obj.amount_paid

                    if final_amount_to_settle != 0:
                        Payment.objects.create(
                            rental=transaction_obj,
                            amount=final_amount_to_settle,
                            payment_method=final_payment_method,
                            payment_type=(
                                PaymentType.RENTAL_FEE
                                if final_amount_to_settle > 0
                                else PaymentType.REFUNDED
                            ),  # Simplified
                            payment_slip=final_payment_slip,
                            notes=f"Final settlement. Fine: {fine_description if fine_description else 'N/A'}. Return notes: {notes_on_return if notes_on_return else 'N/A'}",
                        )

                    transaction_obj.payment_status = PaymentStatus.COMPLETED
                    transaction_obj.save()

                    for (
                        item_detail
                    ) in (
                        transaction_obj.rentalitemdetail_set.all()
                    ):  # Covers both individual and set items
                        item_detail.item.status = (
                            ItemStatus.AVAILABLE
                        )  # Add logic for 'UNDER_MAINTENANCE' based on notes_on_return
                        item_detail.item.save()

                    messages.success(
                        request, f"Rental #{transaction_obj.id} processed and settled successfully."
                    )
                    return redirect("manage_returns")
            except Exception as e:
                messages.error(request, f"An error occurred during settlement: {e}")
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = ReturnSettlementForm()

    context = {
        "transaction": transaction_obj,
        "form": form,
        "initial_balance_due": initial_balance_due,
        "other_payments_made": other_payments_made,
        "title": f"Process Return for Rental #{transaction_obj.id}",
    }
    return render(request, "core/dashboard/pages/process_return.html", context)


@login_required
def booking_history_view(request):
    """Displays the rental transaction history for the logged-in customer."""
    # Ensure the user has a customer profile
    customer = get_object_or_404(Customer, user=request.user)

    # Fetch all rental transactions for this customer, ordered by start date descending
    transactions = RentalTransaction.objects.filter(customer=customer).order_by(
        "-start_date", "-created_at"
    )

    context = {
        "transactions": transactions,
    }
    return render(request, "core/clients/pages/booking_history.html", context)


def manage_products(request):
    context = {
        "delete_confirm_msg": "Are you sure you want to delete this product?",
    }
    return render(request, "core/dashboard/pages/manage-products.html", context)


# -------- Export Sections
#
@login_required
def export_users_excel(request):
    """Exports user data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    users = Account.objects.all().order_by("-date_joined")

    if query:
        users = users.filter(Q(email__icontains=query))

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    # Define headers
    headers = ["Email", "Date Joined", "Last Login", "Is Active", "Is Staff", "Is Admin"]
    sheet.append(headers)

    # Write data rows
    for user in users:
        sheet.append(
            [
                user.email,
                user.date_joined.strftime("%d/%m/%Y %H:%M") if user.date_joined else "",
                user.last_login.strftime("%d/%m/%Y %H:%M") if user.last_login else "N/A",
                "Yes" if user.is_active else "No",
                "Yes" if user.is_staff else "No",
                "Yes" if user.is_superuser else "No",
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=users_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_employees_excel(request):
    """Exports employee data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    employees = Employee.objects.all().order_by("-updated_at")

    if query:
        employees = employees.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(phone_number__icontains=query)
        )

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    # Define headers
    headers = [
        "First Name",
        "Last Name",
        "Phone Number",
        "Address",
        "Employment Date",
        "User Email",
        "User Status",
    ]
    sheet.append(headers)

    # Write data rows
    for emp in employees:
        sheet.append(
            [
                emp.first_name,
                emp.last_name,
                emp.phone_number,
                emp.address,
                emp.employment_date.strftime("%d/%m/%Y") if emp.employment_date else "",
                emp.user.email if emp.user else "N/A",
                "Active" if emp.user and emp.user.is_active else "Inactive",
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=employees_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_customers_excel(request):
    """Exports customer data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    customers = Customer.objects.all().order_by("-updated_at")

    if query:
        customers = customers.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(phone_number__icontains=query)
        )

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Customers"

    # Define headers
    headers = [
        "First Name",
        "Last Name",
        "Phone Number",
        "Address",
        "User Email",
        "User Status",
    ]
    sheet.append(headers)

    # Write data rows
    for customer in customers:
        sheet.append(
            [
                customer.first_name,
                customer.last_name,
                customer.phone_number,
                customer.address,
                customer.user.email if customer.user else "N/A",
                "Active" if customer.user and customer.user.is_active else "Inactive",
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=customers_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_rental_item_types_excel(request):
    """Exports rental item type data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    rental_item_types = RentalItemType.objects.all().order_by("-updated_at")

    if query:
        rental_item_types = rental_item_types.filter(Q(type_name__icontain=query))

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rental Item Types"

    # Define headers
    headers = [
        "Type Name",
        "Size",
        "Capacity",
        "Rental Price Per Day",
        "Replacement Cost",
        "Description",
    ]
    sheet.append(headers)

    # Write data rows
    for item_type in rental_item_types:
        sheet.append(
            [
                item_type.type_name,
                item_type.size or "N/A",
                item_type.capacity or "N/A",
                item_type.rental_price_per_day,
                item_type.replacement_cost,
                item_type.description or "N/A",
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=rental_item_types_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_rental_items_excel(request):
    """Exports rental item data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    rental_items = RentalItem.objects.all().order_by("-updated_at")

    if query:
        rental_items = rental_items.filter(Q(serial_number__icontains=query))

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rental Items"

    # Define headers
    headers = [
        "Serial Number",
        "Item Type",
        "Status",
        "Purchase Date",
        "Last Inspection Date",
        "Condition Notes",
    ]
    sheet.append(headers)

    # Write data rows
    for item in rental_items:
        sheet.append(
            [
                item.serial_number,
                str(item.item_type),  # Use __str__ method
                item.get_status_display(),
                item.purchase_date.strftime("%d/%m/%Y") if item.purchase_date else "",
                item.last_inspection_date.strftime("%d/%m/%Y") if item.last_inspection_date else "",
                item.condition_notes or "N/A",
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=rental_items_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_accessories_excel(request):
    """Exports accessory data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    accessories = Accessory.objects.all().order_by("-updated_at")

    if query:
        accessories = accessories.filter(Q(accessory_name__icontains=query))

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Accessories"

    # Define headers
    headers = [
        "Accessory Name",
        "Item Type",
        "Standard Quantity",
        "Replacement Cost",
    ]
    sheet.append(headers)

    # Write data rows
    for accessory in accessories:
        sheet.append(
            [
                accessory.accessory_name,
                str(accessory.item_type),  # Use __str__ method
                accessory.standard_quantity,
                accessory.replacement_cost,
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=accessories_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_item_sets_excel(request):
    """Exports item set data to an Excel file, respecting current filters."""
    query = request.GET.get("search", "")
    item_sets = (
        ItemSet.objects.all().order_by("name").prefetch_related("itemsetcomponent_set__item_type")
    )

    if query:
        item_sets = item_sets.filter(Q(name__icontains=query) | Q(description__icontains=query))

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Item Sets"

    # Define headers
    headers = [
        "Set Name",
        "Description",
        "Base Price",
        "Replacement Deposit",
        "Components",
    ]
    sheet.append(headers)

    # Write data rows
    for item_set in item_sets:
        components_list = ", ".join(
            [f"{c.item_type.type_name} x {c.quantity}" for c in item_set.itemsetcomponent_set.all()]
        )
        sheet.append(
            [
                item_set.name,
                item_set.description or "N/A",
                item_set.base_price,
                item_set.replacement_deposit,
                components_list or "None",
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=item_sets_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def export_rental_transactions_excel(request):
    """Exports rental transaction data to an Excel file, respecting current filters."""

    transactions_list = (
        RentalTransaction.objects.select_related("customer__user")
        .prefetch_related(
            "rentalitemdetail_set__item__item_type",
            "rentalsetdetail_set__item_set",
            "payment_set",
        )
        .order_by("-start_date", "-created_at")
    )

    # --- Apply Filtering Logic (same as manage_rental_transactions) ---
    status_filter = request.GET.get("status")
    customer_filter = request.GET.get("customer")
    date_from_filter = request.GET.get("date_from")
    date_to_filter = request.GET.get("date_to")

    if status_filter:
        transactions_list = transactions_list.filter(payment_status=status_filter)

    if customer_filter:
        transactions_list = transactions_list.filter(
            Q(customer__first_name__icontains=customer_filter)
            | Q(customer__last_name__icontains=customer_filter)
            | Q(customer__user__email__icontains=customer_filter)
        )

    if date_from_filter:
        try:
            start_date = timezone.datetime.strptime(date_from_filter, "%Y-%m-%d").date()
            transactions_list = transactions_list.filter(start_date__gte=start_date)
        except ValueError:
            pass  # Ignore invalid date filter for export

    if date_to_filter:
        try:
            end_date = timezone.datetime.strptime(date_to_filter, "%Y-%m-%d").date()
            transactions_list = transactions_list.filter(end_date__lte=end_date)
        except ValueError:
            pass  # Ignore invalid date filter for export

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rental Transactions"

    # Define headers
    headers = [
        "Transaction ID",
        "Customer Name",
        "Customer Email",
        "Rental Dates",
        "Total Days",
        "Items/Sets Rented",
        "Total Deposit",
        "Total Rental Cost",
        "Amount Paid",
        "Payment Status",
        "Created At",
    ]
    sheet.append(headers)

    # Write data rows
    for transaction in transactions_list:
        items_sets_list = "; ".join(
            [
                f"{d.quantity} x {d.item_set.name} (Set)"
                for d in transaction.rentalsetdetail_set.all()
            ]
            + [
                f"{d.quantity} x {d.item.item_type} ({d.item.serial_number})"
                for d in transaction.rentalitemdetail_set.filter(set_rental__isnull=True)
            ]
        )
        sheet.append(
            [
                transaction.id,
                transaction.customer.first_name + " " + transaction.customer.last_name,
                transaction.customer.user.email if transaction.customer.user else "N/A",
                f"{transaction.start_date.strftime('%d/%m/%Y')} - {transaction.end_date.strftime('%d/%m/%Y')}",
                transaction.total_rental_days,
                items_sets_list or "None",
                transaction.total_deposit,
                transaction.total_rental_cost,
                transaction.amount_paid,
                transaction.get_payment_status_display(),
                transaction.created_at.strftime("%d/%m/%Y %H:%M"),
            ]
        )

    # Optional: Adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_width = True

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=rental_transactions_export.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response
